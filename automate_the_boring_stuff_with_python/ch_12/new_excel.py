import openpyxl

wb = openpyxl.Workbook()
print(wb.get_sheet_names())

ws = wb.active
print(ws.title)

ws.title = 'Omlet'
print(wb.get_sheet_names())

wb.create_sheet()
print(wb.get_sheet_names())
wb.create_sheet(index=0 , title='1st sheet')
print(wb.get_sheet_names())
wb.create_sheet(index=2 , title='Middle sheet')
print(wb.get_sheet_names())

print(wb.get_sheet_by_name('1st sheet'))

test = {'A':1 , 'B':2 , 'C':3}
ws = wb.get_sheet_by_name('1st sheet')
ws['A1'] = test

wb.save('example.xlsx')

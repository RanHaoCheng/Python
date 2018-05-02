import sys
import New
import Original
import openpyxl


def saveResultToExcel(newList , oriList):
    # wb = openpyxl.Workbook()
    # sheetList = wb.get_sheet_names()
    # ws = wb.get_sheet_by_name(sheetList[0])
    # rowIndex = 1
    # ws['B'+str(rowIndex)].value = 'RD BOM Data'
    # ws['C'+str(rowIndex)].value = 'R-BOM  Data'
    # rowIndex = rowIndex+1
    # colIndex = 1    # get_column_letter()

    length = max(len(newList), len(oriList))
    if len(newList) == length :
        for i in range(0,length):
            isSameNumber = False
            data = newList[i]['NUMBER']
            for j in range(0,len(oriList)):
                if(data in oriList[j].values()):
                    isSameNumber = True

                    print('Diff datas')
                    print('\t\tOriginal Data\t\t\tNew Data')
                    print('NAME\t\t'+oriList[j]['NAME']+'\t\t'+newList[i]['NAME'])
                    print('NUMBER\t\t'+oriList[j]['NUMBER']+'\t\t'+newList[i]['NUMBER'])
                    print('PLACEMENT\t'+oriList[j]['PLACEMENT']+'\t\t\t\t'+newList[i]['PLACEMENT'])
                    print('REPLACEMENT\t'+oriList[j]['REPLACEMENT']+'\t\t\t\t'+newList[i]['REPLACEMENT'])
                    print('USAGE\t\t'+str(oriList[j]['USAGE'])+'\t\t\t\t'+str(newList[i]['USAGE']))
                    print('SPEC\t\t'+oriList[j]['SPEC']+'\t\t'+newList[i]['SPEC'])
                    print('')
            if isSameNumber == False:
                print('Datas that new BOM have')
                print('NAME\t\t'+newList[i]['NAME'])
                print('NUMBER\t\t'+newList[i]['NUMBER'])
                print('PLACEMENT\t'+newList[i]['PLACEMENT'])
                print('REPLACEMENT\t'+newList[i]['REPLACEMENT'])
                print('USAGE\t\t'+str(newList[i]['USAGE']))
                print('SPEC\t\t'+newList[i]['SPEC'])
                print('')
                

    else:           
        for i in range(0,length):
            isSameNumber = False
            data = oriList[i]['NUMBER']
            for j in range(0,len(newIndex)):
                if(data in newIndex[j].values()):
                    print('\t\tOriginal Data\t\t\tNew Data')
                    print('NAME\t\t'+oriList[j]['NAME']+'\t\t'+newList[i]['NAME'])
                    print('NUMBER\t\t'+oriList[j]['NUMBER']+'\t\t'+newList[i]['NUMBER'])
                    print('PLACEMENT\t'+oriList[j]['PLACEMENT']+'\t\t\t\t'+newList[i]['PLACEMENT'])
                    print('REPLACEMENT\t'+oriList[j]['REPLACEMENT']+'\t\t\t\t'+newList[i]['REPLACEMENT'])
                    print('USAGE\t\t'+str(oriList[j]['USAGE'])+'\t\t\t\t'+str(newList[i]['USAGE']))
                    print('SPEC\t\t'+oriList[j]['SPEC']+'\t\t'+newList[i]['SPEC'])
                    print('')
            if isSameNumber == False:
                print('Datas that original BOM have')
                print('NAME\t\t'+oriList[i]['NAME'])
                print('NUMBER\t\t'+oriList[i]['NUMBER'])
                print('PLACEMENT\t'+oriList[i]['PLACEMENT'])
                print('REPLACEMENT\t'+oriList[i]['REPLACEMENT'])
                print('USAGE\t\t'+str(oriList[i]['USAGE']))
                print('SPEC\t\t'+oriList[i]['SPEC'])
                print('')

    # wb.save('CompareResult.xlsx')

def main():
    NewBOMList = New.New
    OriginalBOMList = Original.Original

    # print(len(NewBOMList))
    # print(len(OriginalBOMList))

    #tmp = NewBOMList.pop(1)
    
    oriIndex = []
    newIndex = []

    #print(len(NewBOMList))
    # tmp = NewBOMList[1]['NUMBER']
    # print(tmp)
    # for i in range(0,len(OriginalBOMList)):
    #     if(tmp in OriginalBOMList[i].values()):
    #         print(str(i) )

    # print('Component in original BOM but not in new :')
    for i in range(0,len(OriginalBOMList)):
        BOMdata = OriginalBOMList[i]
        try:
            NewBOMList.index(BOMdata)
        except ValueError:
            # print('Index : ' + str(7+i) + ' Component number ' + BOMdata['NUMBER'])
            oriIndex.append(BOMdata)

    # print(oriIndex)

    # print('\nComponent in new BOM but not in original :')
    for i in range(0,len(NewBOMList)):
        BOMdata = NewBOMList[i]
        try:
            OriginalBOMList.index(BOMdata)
        except ValueError:
            newIndex.append(BOMdata)
            # print('Index : ' + str(7+i) + ' Component number ' + BOMdata['NUMBER'])

    # print(newIndex)

    # length = max(len(newIndex), len(oriIndex))
    # if len(newIndex) == length :
    #     for i in range(0,length):
    #         isSameNumber = False
    #         data = newIndex[i]['NUMBER']
    #         for j in range(0,len(oriIndex)):
    #             if(data in oriIndex[j].values()):
    #                 isSameNumber = True
    #                 print('Diff datas')
    #                 print('\t\tOriginal Data\t\t\tNew Data')
    #                 print('NAME\t\t'+oriIndex[j]['NAME']+'\t\t'+newIndex[i]['NAME'])
    #                 print('NUMBER\t\t'+oriIndex[j]['NUMBER']+'\t\t'+newIndex[i]['NUMBER'])
    #                 print('PLACEMENT\t\t'+oriIndex[j]['PLACEMENT']+'\t\t'+newIndex[i]['PLACEMENT'])
    #                 print('REPLACEMENT\t\t'+oriIndex[j]['REPLACEMENT']+'\t\t'+newIndex[i]['REPLACEMENT'])
    #                 print('USAGE\t\t'+str(oriIndex[j]['USAGE'])+'\t\t'+str(newIndex[i]['USAGE']))
    #                 print('SPEC\t\t'+oriIndex[j]['SPEC']+'\t\t'+newIndex[i]['SPEC'])
    #                 print('')
    #         if isSameNumber == False:
    #             print('New Datas')
    #             print('NAME\t\t'+newIndex[i]['NAME'])
    #             print('NUMBER\t\t'+newIndex[i]['NUMBER'])
    #             print('PLACEMENT\t\t'+newIndex[i]['PLACEMENT'])
    #             print('REPLACEMENT\t\t'+newIndex[i]['REPLACEMENT'])
    #             print('USAGE\t\t'+str(newIndex[i]['USAGE']))
    #             print('SPEC\t\t'+newIndex[i]['SPEC'])
    #             print('')
                

    # else:           
    #     for i in range(0,length):
    #         isSameNumber = False
    #         data = oriIndex[i]['NUMBER']
    #         for j in range(0,len(newIndex)):
    #             if(data in newIndex[j].values()):
    #                 print('\t\tOriginal Data\t\t\tNew Data')
    #                 print('NAME\t\t'+oriIndex[j]['NAME']+'\t\t'+newIndex[i]['NAME'])
    #                 print('NUMBER\t\t'+oriIndex[j]['NUMBER']+'\t\t'+newIndex[i]['NUMBER'])
    #                 print('PLACEMENT\t\t'+oriIndex[j]['PLACEMENT']+'\t\t'+newIndex[i]['PLACEMENT'])
    #                 print('REPLACEMENT\t\t'+oriIndex[j]['REPLACEMENT']+'\t\t'+newIndex[i]['REPLACEMENT'])
    #                 print('USAGE\t\t'+str(oriIndex[j]['USAGE'])+'\t\t'+str(newIndex[i]['USAGE']))
    #                 print('SPEC\t\t'+oriIndex[j]['SPEC']+'\t\t'+newIndex[i]['SPEC'])
    #                 print('')
    #         if isSameNumber == False:
    #             print('New Datas')
    #             print('NAME\t\t'+oriIndex[i]['NAME'])
    #             print('NUMBER\t\t'+oriIndex[i]['NUMBER'])
    #             print('PLACEMENT\t\t'+oriIndex[i]['PLACEMENT'])
    #             print('REPLACEMENT\t\t'+oriIndex[i]['REPLACEMENT'])
    #             print('USAGE\t\t'+str(oriIndex[i]['USAGE']))
    #             print('SPEC\t\t'+oriIndex[i]['SPEC'])
    #             print('')
                    
    
    saveResultToExcel(newIndex , oriIndex)

if __name__ == "__main__":
    main()


import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
ws = wb.get_sheet_by_name('Sheet')

# updates new price in dict here
PRICE_UPDATES =  {
        'Garlic':3.07,
        'Celery':1.19,
        'Lemon':1.27
        }

# Loop to update prices
for rows in range(2,ws.max_row):
    produceName = ws.cell(row=rows , column=1).value
    if produceName in PRICE_UPDATES:
        ws.cell(row=rows , column=2).value = PRICE_UPDATES[produceName]


wb.save('produceSales.xlsx')


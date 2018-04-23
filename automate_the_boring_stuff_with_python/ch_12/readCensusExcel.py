#! python3

import openpyxl , pprint
print("Opening workbook...")

#workbook
wb = openpyxl.load_workbook('censuspopdata.xlsx')

#worksheet
ws = wb.get_sheet_by_name('Population by Census Tract')
countyData = { }

print('Reading rows...')
for row in range(2 , ws.max_row+1):
    state = ws['B'+str(row)].value
    county = ws['C'+str(row)].value
    pop  = ws['D'+str(row)]

    countyData.setdefault(state , {})
    countyData[state].setdefault(county , {'tracts' : 0 , 'pop' : 0})

    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop.value)

print('Writing results...')
resultFile = open('census2010.py' , 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')


import os
import census2010
census2010.allData['AK']['Anchorage']
anchoragePop = census2010.allData['AK']['Anchorage']['pop']
print('The 2010 population of Anchorage was ' + str(anchoragePop))

#!/usr/bin/python
# -*- coding: UTF-8 -*-

import openpyxl

wb = openpyxl.load_workbook('RDPIO301XX1421-ELE-V1.0_20180208_Jessie.xlsx')
sheetList = wb.get_sheet_names()
print(sheetList)

sheet = wb.get_sheet_by_name(sheetList[0])
print(sheet)
print(type(sheet))

print(sheet.title)
anotherSheet = wb.active
print(anotherSheet)

print(sheet['A7'])
print(sheet['A7'].value)
c=sheet['B7']
print(c.value)

c_r= 'Row ' + str(c.row) + ' , Column ' + c.column + ' is ' + c.value
print(c_r)
print(sheet['C7'].value)

print(sheet.cell(row=7,column=2))
print(sheet.cell(row=7,column=2).value)

#for i in range(1,20):
#    print(i , sheet.cell(row=i , column=1).value)

print("Sheet max row : " + str(sheet.max_row))
print("Sheet max column : " + str(sheet.max_column))

from openpyxl.utils import get_column_letter , column_index_from_string

print(get_column_letter(900))
print(column_index_from_string('AA'))


tuple(sheet['A7':'G20'])

for rowOfCellObjs in sheet['A7':'G20']:
    for cellObj in rowOfCellObjs:
        print(cellObj.coordinate , cellObj.value , type(cellObj.value))
        #print(type(cellObj.value))
    print('---END OF ROW---')

sheet = wb.active

for colObjs in sheet.iter_cols(min_row=7 , max_row=sheet.max_row , min_col=1 , max_col=2):
    for colCell in colObjs:
        print(colCell , colCell.value)
print()
for rowObjs in sheet.iter_rows(min_row=7 , max_row=sheet.max_row , max_col=3):
    for rowCells in rowObjs:
        print(rowCells , rowCells.value)


